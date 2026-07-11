"""
-----------------------------------
File: gedcom_analysis.py

Summary: Parses a GEDCOM file and exports all individuals and family
         relationships to both CSV and a formatted Excel workbook.
         Captures: names, sex, birth/death/burial/marriage dates and places,
         parents, spouses, and children.
         Skips: sources, images, notes, and other non-essential records.

Architect & Designer: Andy Askey
Coder (AI Assistant): Anthropic Claude

License: Apache License 2.0
-----------------------------------
"""

import argparse
import csv
import os
import re
import sys, json
from collections import Counter

from networkx.generators.geometric import geographical_threshold_graph
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pandas.core.dtypes.common import is_numeric_dtype

import common_utils
import gedcom_to_csv
import gen_logging

# ==============================================================================
# CONFIGURATION — edit these paths before running
# ==============================================================================
OUTPUT_DIR = r"E:\Users\Andy\PycharmProjects\Genealogy\output"
INPUT_GEDCOM = r"C:\tempc\ShortTermCSVfiles\test_census.ged"


# ==============================================================================

def add_place(p):
    p1 = p.lower().strip()
    places.append(p1)


def parse_gedcom(filepath):
    individuals = {}
    families = {}

    cur_indi_id = None
    cur_fam_id = None
    cur_event = None

    WANTED_EVENTS = {'BIRT', 'DEAT', 'BURI', 'MARR', 'CENS', 'RESI', 'PAGE'}

    def new_indi():
        return {
            'id': '', 'first_name': '', 'last_name': '', 'sex': '',
            'birth_date': '', 'birth_place': '',
            'death_date': '', 'death_place': '',
            'burial_date': '', 'burial_place': '',
            'fams': [],
            'famc': [],
            'events': [],
        }

    def new_fam():
        return {
            'id': '', 'husb': '', 'wife': '',
            'marr_date': '', 'marr_place': '',
            'children': [],
        }

    linecnt = 0
    with open(filepath, encoding='utf-8', errors='replace') as f:
        for raw_line in f:
            line = raw_line.rstrip('\r\n')

            linecnt += 1

            if not line.strip():
                continue

            parts = line.split(' ', 2)
            if len(parts) < 2:
                continue

            level = parts[0].strip()
            tag = parts[1].strip() if len(parts) > 1 else ''
            value = parts[2].strip() if len(parts) > 2 else ''

            if level == '0':
                cur_event = None
                if tag.startswith('@I') and value == 'INDI':
                    cur_indi_id = tag.strip('@')
                    cur_fam_id = None
                    individuals[cur_indi_id] = new_indi()
                    individuals[cur_indi_id]['id'] = cur_indi_id
                elif tag.startswith('@F') and value == 'FAM':
                    cur_fam_id = tag.strip('@')
                    cur_indi_id = None
                    families[cur_fam_id] = new_fam()
                    families[cur_fam_id]['id'] = cur_fam_id
                else:
                    cur_indi_id = None
                    cur_fam_id = None
                continue

            if cur_indi_id:
                indi = individuals[cur_indi_id]
                if level == '1':
                    cur_event = None
                    if tag == 'NAME':
                        slash_parts = value.split('/')
                        if len(slash_parts) >= 2:
                            indi['last_name'] = slash_parts[1].strip()
                            indi['first_name'] = slash_parts[0].strip()
                        else:
                            indi['first_name'] = value.replace('/', ' ').strip()
                        continue
                    if tag == 'SEX':
                        indi['sex'] = value
                        continue
                    if tag == 'FAMS':
                        indi['fams'].append(value.strip('@'))
                        continue
                    if tag == 'FAMC':
                        indi['famc'].append(value.strip('@'))
                        continue
                    if tag in WANTED_EVENTS:
                        cur_event = tag
                        if cur_event in ('CENS', 'RESI'):
                            indi['events'].append({'date': '', 'place': ''})
                        continue

                if level == '2' and cur_event:
                    if tag == 'DATE':
                        if cur_event == 'BIRT':
                            indi['birth_date'] = value
                        elif cur_event == 'DEAT':
                            indi['death_date'] = value
                        elif cur_event == 'BURI':
                            indi['burial_date'] = value
                        elif cur_event in ('CENS', 'RESI') and indi['events']:
                            indi['events'][-1]['date'] = value
                    elif tag == 'PLAC':
                        if cur_event == 'BIRT':
                            indi['birth_place'] = value
                        elif cur_event == 'DEAT':
                            indi['death_place'] = value
                        elif cur_event == 'BURI':
                            indi['burial_place'] = value
                        elif cur_event in ('CENS', 'RESI') and indi['events']:
                            indi['events'][-1]['place'] = value
                            add_place(common_utils.extract_state(value))
                    continue

                if level == '3':
                    if tag in ('PAGE'):
                        if "census" in line.lower():
                            year_pattern = re.compile(r"(\d{4})\s*U\s*S\s*Census|Year:\s*(\d{4})", re.IGNORECASE)
                            place_pattern = re.compile(r"Census\s+Place:\s*([^;]+)")

                            year_match = year_pattern.search(line)
                            place_match = place_pattern.search(line)

                            year = "Not Found"
                            if year_match:
                                year = year_match.group(1) if year_match.group(1) else year_match.group(2)

                            place = place_match.group(1).strip() if place_match else "Not Found"

                            if year != "Not Found" or place != "Not Found":
                                indi['events'].append({'date': year, 'place': place})
                                add_place(common_utils.extract_state(place))
                                logger.info(f"{linecnt}  {line}")

            if cur_fam_id:
                fam = families[cur_fam_id]
                if level == '1':
                    cur_event = None
                    if tag == 'HUSB':
                        fam['husb'] = value.strip('@')
                        continue
                    if tag == 'WIFE':
                        fam['wife'] = value.strip('@')
                        continue
                    if tag == 'CHIL':
                        fam['children'].append(value.strip('@'))
                        continue
                    if tag == 'MARR':
                        cur_event = 'MARR'
                        continue

                if level == '2' and cur_event == 'MARR':
                    if tag == 'DATE':
                        fam['marr_date'] = value
                    elif tag == 'PLAC':
                        fam['marr_place'] = value
                    continue

    return individuals, families


def build_individuals_rows(individuals, families):
    rows = []

    def name_of(indi_id):
        if not indi_id or indi_id not in individuals:
            return ''
        i = individuals[indi_id]
        return f"{i['first_name']} {i['last_name']}".strip()

    def birth_year_of(indi_id):
        if not indi_id or indi_id not in individuals:
            return ''
        date_str = individuals[indi_id].get('birth_date', '')
        m = re.search(r'\b(1[0-9]{3}|20[0-2][0-9])\b', date_str)
        return m.group(1) if m else ''

    def birth_place_of(indi_id):
        if not indi_id or indi_id not in individuals:
            return ''
        return common_utils.extract_state(individuals[indi_id].get('birth_place', ''))

    def birth_county_of(indi_id):
        if not indi_id or indi_id not in individuals:
            return ''
        return common_utils.extract_county(individuals[indi_id].get('birth_place', ''))

    for indi_id, i in individuals.items():
        if not i.get('first_name') or '-' in i.get('first_name') or '-' in i.get('last_name') or 'Living' in i.get(
                'first_name'):
            continue

        father_id = ''
        father_name = ''
        father_birth_year = ''
        father_birth_place = ''
        father_birth_county = ''
        mother_id = ''
        mother_name = ''
        mother_birth_year = ''
        mother_birth_place = ''
        mother_birth_county = ''
        if i['famc']:
            fam = families.get(i['famc'][0])
            if fam:
                father_id = fam['husb']
                father_name = name_of(fam['husb'])
                father_birth_year = birth_year_of(fam['husb'])
                father_birth_place = common_utils.extract_state(birth_place_of(fam['husb']))
                father_birth_county = birth_county_of(fam['husb'])
                mother_id = fam['wife']
                mother_name = name_of(fam['wife'])
                mother_birth_year = birth_year_of(fam['wife'])
                mother_birth_place = common_utils.extract_state(birth_place_of(fam['wife']))
                mother_birth_county = birth_county_of(fam['wife'])
                add_place(common_utils.extract_state(mother_birth_place))
                add_place(common_utils.extract_state(father_birth_place))

        spouse_ids = []
        spouse_names = []
        spouse_birth_years = []
        spouse_birth_places = []
        marr_dates = []
        marr_places = []
        spouse_birth_counties = []
        for fam_id in i['fams']:
            fam = families.get(fam_id)
            if not fam:
                continue
            spouse_id = fam['wife'] if fam['husb'] == indi_id else fam['husb']
            spouse_ids.append(spouse_id)
            spouse_names.append(name_of(spouse_id))
            spouse_birth_years.append(birth_year_of(spouse_id))
            spouse_birth_places.append(common_utils.extract_state(birth_place_of(spouse_id)))
            spouse_birth_counties.append(birth_county_of(spouse_id))
            marr_dates.append(fam['marr_date'])
            marr_places.append(fam['marr_place'])

        child_ids = []
        child_names = []
        child_byrs = []
        for fam_id in i['fams']:
            fam = families.get(fam_id)
            if fam:
                for c_id in fam['children']:
                    child_ids.append(c_id)
                    child_names.append(name_of(c_id))
                    child_byrs.append(birth_year_of(c_id))

        residences = {str(yr): '' for yr in range(1850, 1960, 10)}
        for ev in i['events']:
            for yr in residences:
                if yr in ev['date']:
                    residences[yr] = ev['place']

        residences_extracted = {}
        for yr in range(1850, 1960, 10):
            residences_extracted[f"{yr} County"] = common_utils.extract_county(residences[str(yr)])
            residences_extracted[f"{yr} State"] = common_utils.extract_state(residences[str(yr)])

        rows.append({
            'ID': indi_id,
            'First Name': i['first_name'],
            'Last Name': i['last_name'],
            'Sex': i['sex'],
            'Birth Date': i['birth_date'],
            'Birth Place': i['birth_place'],
            'Birth County': common_utils.extract_county(i['birth_place']),
            'Death Date': i['death_date'],
            'Death Place': i['death_place'],
            'Burial Date': i['burial_date'],
            'Burial Place': i['burial_place'],
            '1850 Place': residences['1850'],
            '1860 Place': residences['1860'],
            '1870 Place': residences['1870'],
            '1880 Place': residences['1880'],
            '1890 Place': residences['1890'],
            '1900 Place': residences['1900'],
            '1910 Place': residences['1910'],
            '1920 Place': residences['1920'],
            '1930 Place': residences['1930'],
            '1940 Place': residences['1940'],
            '1950 Place': residences['1950'],
            **residences_extracted,
            'Father ID': father_id,
            'Father': father_name,
            'Father Birth Year': father_birth_year,
            'Father Birth Place': father_birth_place,
            'Father Birth County': father_birth_county,
            'Mother ID': mother_id,
            'Mother': mother_name,
            'Mother Birth Year': mother_birth_year,
            'Mother Birth Place': mother_birth_place,
            'Mother Birth County': mother_birth_county,
            'Spouse ID(s)': ' | '.join(filter(None, spouse_ids)),
            'Spouse(s)': ' | '.join(filter(None, spouse_names)),
            'Spouse Birth Year(s)': ' | '.join(filter(None, spouse_birth_years)),
            'Spouse Birth Place(s)': ' | '.join(filter(None, spouse_birth_places)),
            'Spouse Birth County(s)': ' | '.join(filter(None, spouse_birth_counties)),
            'Marriage Date(s)': ' | '.join(filter(None, marr_dates)),
            'Marriage Place(s)': ' | '.join(filter(None, marr_places)),
            'Num Children': len(child_ids),
            'Children ID(s)': ' | '.join(filter(None, child_ids)),
            'Children': ' | '.join(filter(None, child_names)),
            'Children Birth Yr': ' | '.join(filter(None, child_byrs)),
            'Matched': ''
        })

    rows.sort(key=lambda r: (r['Last Name'].upper(), r['First Name'].upper()))
    return rows


def build_families_rows(individuals, families):
    rows = []

    def name_of(indi_id):
        if not indi_id or indi_id not in individuals:
            return ''
        i = individuals[indi_id]
        return f"{i['first_name']} {i['last_name']}".strip()

    def dates_of(indi_id, field):
        if not indi_id or indi_id not in individuals:
            return ''
        return individuals[indi_id].get(field, '')

    for fam_id, fam in families.items():
        child_names = [name_of(c) for c in fam['children']]
        rows.append({
            'Family ID': fam_id,
            'Husband ID': fam['husb'],
            'Husband': name_of(fam['husb']),
            'Husb Birth': dates_of(fam['husb'], 'birth_date'),
            'Husb Death': dates_of(fam['husb'], 'death_date'),
            'Wife ID': fam['wife'],
            'Wife': name_of(fam['wife']),
            'Wife Birth': dates_of(fam['wife'], 'birth_date'),
            'Wife Death': dates_of(fam['wife'], 'death_date'),
            'Marriage Date': fam['marr_date'],
            'Marriage Place': fam['marr_place'],
            'Num Children': len(fam['children']),
            'Children ID(s)': ' | '.join(filter(None, fam['children'])),
            'Children': ' | '.join(filter(None, child_names)),
        })

    rows.sort(key=lambda r: r['Husband'].upper())
    return rows


def analyze_homeland_counties(individuals, families, logger, out_dir):
    """Extracts and counts all counties in the GEDCOM to find the family's 'Homeland' regions."""
    county_counter = Counter()

    def add_county(plac_str):
        if not plac_str: return
        c = common_utils.extract_county(plac_str)
        if c:
            c_clean = c.upper().replace(" COUNTY", "").replace(" CO.", "").strip()
            if c_clean and len(c_clean) > 2:
                county_counter[c_clean] += 1

    for i in individuals.values():
        add_county(i.get('birth_place'))
        add_county(i.get('death_place'))
        add_county(i.get('burial_place'))
        for ev in i.get('events', []):
            add_county(ev.get('place'))

    for f in families.values():
        add_county(f.get('marr_place'))

    top_counties = county_counter.most_common()

    logger.info("\n==================================================")
    logger.info("   GEDCOM 'HOMELAND' COUNTY STATISTICAL ANALYSIS")
    logger.info("==================================================")
    for county, count in top_counties[:25]:
        logger.info(f"  {county:<20} : {count:,} records")
    logger.info("==================================================\n")

    homeland_file = os.path.join(out_dir, "homeland_counties.json")
    with open(homeland_file, "w", encoding="utf-8") as f:
        json.dump(dict(top_counties), f, indent=4)

    logger.info(f"Homeland analysis exported to: {homeland_file}")


def parse_date(date_str):
    """Extracts a 4-digit year and the month number from a date string."""
    if not date_str:
        return "", ""
    date_str = str(date_str).upper()
    byr_match = re.search(r'\b(1[456789]\d\d|20\d\d)\b', date_str)
    byr = byr_match.group(1) if byr_match else ""

    bmo = ""
    months = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
    for i, m in enumerate(months, 1):
        if m in date_str:
            bmo = str(i)
            break
    return byr, bmo


def get_parent_bpl(indi_id, parent_type, individuals, families):
    if not indi_id or indi_id not in individuals: return ""
    i = individuals[indi_id]
    if not i['famc']: return ""
    fam = families.get(i['famc'][0])
    if not fam: return ""
    p_id = fam['husb'] if parent_type == 'father' else fam['wife']
    if not p_id or p_id not in individuals: return ""
    return individuals[p_id].get('birth_place', '')


def writeSQL(first, last, byr, bpl, mbpl, fbpl):
    tmp = common_utils.get_bpl_prefixes(bpl)
    bplcode = int(tmp[0]) if tmp else 0
    tmp = common_utils.get_bpl_prefixes(mbpl)
    mbplcode = int(tmp[0]) if tmp else 0
    tmp = common_utils.get_bpl_prefixes(fbpl)
    fbplcode = int(tmp[0]) if tmp else 0

    byr_int = int(byr) if byr and str(byr).isnumeric() else 0

    sql_query = f"""SELECT 
    histid, 
    first_name, 
    last_name, 
    sex, 
    birthyr, 
    bpld, 
    fbpl, 
    mbpl 
FROM 
    individuals 
WHERE 
    CAST(birthyr AS INTEGER) BETWEEN {byr_int - 5} AND {byr_int + 5} 
    AND (CASE WHEN CAST(bpld AS INTEGER) >= 1000 THEN CAST(bpld AS INTEGER) / 100 ELSE CAST(bpld AS INTEGER) END) = {bplcode} 
    AND (CASE WHEN CAST(fbpl AS INTEGER) >= 1000 THEN CAST(fbpl AS INTEGER) / 100 ELSE CAST(fbpl AS INTEGER) END) = {fbplcode} 
    AND (CASE WHEN CAST(mbpl AS INTEGER) >= 1000 THEN CAST(mbpl AS INTEGER) / 100 ELSE CAST(mbpl AS INTEGER) END) = {mbplcode};"""

    fname = f"../../sql/{first}_{last}_{byr}.sql"
    with open(fname, 'w', encoding='utf-8') as f:
        f.write(sql_query)


def to_pos_int(str):
    if not str: return 0
    if not str.isnumeric(): return 0
    num = int(str)
    if num < 1: num = 0
    return num


def to_len(str):
    if not str: return 0
    if str.isnumeric(): return 0
    num = len(str.strip())
    if num < 1: num = 0
    return num


def build_couples_rows(individuals, families):
    """Generates the Expanded Nuclear Family Fingerprint."""
    records = []
    seen_couples = set()

    def get_res_c(indi_id, year_str):
        if not indi_id or indi_id not in individuals: return ""
        for ev in individuals[indi_id].get('events', []):
            if year_str in ev['date']:
                cty = common_utils.extract_county(ev['place'])
                return cty
        return ""

    def get_res_s(indi_id, year_str):
        if not indi_id or indi_id not in individuals: return ""
        for ev in individuals[indi_id].get('events', []):
            if year_str in ev['date']:
                st = common_utils.extract_state(ev['place'])
                return st
        return ""

    def get_fingerprint(indi_id):
        if not indi_id or indi_id not in individuals:
            return "", "", "", "", "", "", "", {str(yr): "" for yr in range(1850, 1960, 10)}, {str(yr): "" for yr in
                                                                                               range(1850, 1960,
                                                                                                     10)}, "", ""
        i = individuals[indi_id]
        first, last = i.get('first_name', ''), i.get('last_name', '')
        byr, bmo = parse_date(i.get('birth_date', ''))
        dyr, _ = parse_date(i.get('death_date', ''))

        bpl = common_utils.extract_state(i.get('birth_place', ''))
        dpl = common_utils.extract_state(i.get('death_place', ''))

        fbpl = common_utils.extract_state(get_parent_bpl(indi_id, 'father', individuals, families))
        mbpl = common_utils.extract_state(get_parent_bpl(indi_id, 'mother', individuals, families))

        residences_st = {str(yr): get_res_s(indi_id, str(yr)) for yr in range(1850, 1960, 10)}
        residences_cty = {str(yr): get_res_c(indi_id, str(yr)) for yr in range(1850, 1960, 10)}

        return first, last, byr, bmo, bpl, fbpl, mbpl, residences_st, residences_cty, dyr, dpl

    # Process Nuclear Families
    for fam_id, fam in families.items():
        h_id, w_id = fam.get('husb'), fam.get('wife')
        h_first, h_last, h_byr, h_bmo, h_bpl, h_fbpl, h_mbpl, h_st, h_cty, h_dyr, h_dpl = get_fingerprint(h_id)
        w_first, w_last, w_byr, w_bmo, w_bpl, w_fbpl, w_mbpl, w_st, w_cty, w_dyr, w_dpl = get_fingerprint(w_id)

        if to_pos_int(h_dyr) < 1850 or to_pos_int(w_dyr) < 1850:
            continue
        if to_len(h_first) < 1 or to_len(w_first) < 1 or to_len(h_last) < 1 or to_len(w_last) < 1:
            continue
        if to_len(h_bpl) < 1 or to_len(w_bpl) < 1 or to_len(w_fbpl) < 1 or to_len(w_last) < 1:
            continue

        # Filter out fake names but ALLOW families without a wife's first name
        if not h_first or not h_last:
            logger.info(f"Skipping {fam_id}: Missing Husband Name ({h_first} {h_last})")
            continue

        marr_yr, _ = parse_date(fam.get('marr_date', ''))
        marr_pl = common_utils.extract_state(fam.get('marr_place', ''))
        num_children = len(fam.get('children', []))

        kid_fingerprint = 0
        child_byrs = []
        for c_id in fam.get('children', []):
            c_byr, _ = parse_date(individuals.get(c_id, {}).get('birth_date', ''))
            if c_byr:
                child_byrs.append(c_byr)
                kid_fingerprint += common_utils.safe_cast(c_byr)

        # kid_fingerprint = 0
        # kidbys = '|'.join(sorted(child_byrs))
        # for k in child_byrs:
        #     kid_fingerprint += int(k)

        if ('--' in h_first or ' --' in h_last or 'Living' in h_last or '[' in h_last) or \
                ('--' in w_last or 'Living' in w_last or '[' in w_last):
            logger.info(f"Skipping {fam_id}: Fake/Living Name Detected")
            continue

        gtg = False
        if h_byr and h_byr.isnumeric() and int(h_byr) < 1950:
            gtg = True
        if not gtg: continue

        couple_key = fam_id
        if couple_key not in seen_couples:

            row = {
                'h_first': h_first, 'h_last': h_last, 'h_byr': h_byr, 'h_bmo': h_bmo,
                'h_bpl': h_bpl, 'h_fbpl': h_fbpl, 'h_mbpl': h_mbpl,
            }
            for yr in range(1850, 1960, 10):
                row[f'h_rsst_{yr}'] = h_st[str(yr)]
                row[f'h_rscty_{yr}'] = h_cty[str(yr)]
            row.update({
                'h_dyr': h_dyr, 'h_dpl': h_dpl,
                'w_first': w_first, 'w_last': w_last, 'w_byr': w_byr, 'w_bmo': w_bmo,
                'w_bpl': w_bpl, 'w_fbpl': w_fbpl, 'w_mbpl': w_mbpl,
            })
            for yr in range(1850, 1960, 10):
                row[f'w_rsst_{yr}'] = w_st[str(yr)]
                row[f'w_rscty_{yr}'] = w_cty[str(yr)]
            row.update({
                'w_dyr': w_dyr, 'w_dpl': w_dpl,
                'marr_yr': marr_yr, 'marr_pl': marr_pl, 'num_children': num_children,
                'kid_fingerprint': kid_fingerprint
            })
            records.append(row)
            seen_couples.add(couple_key)

    # Process Lone Wolves (Unmarried individuals)
    for indi_id, i in individuals.items():
        if not i.get('fams'):
            first, last, byr, bmo, bpl, fbpl, mbpl, res, res_cty, dyr, dpl = get_fingerprint(indi_id)
            if not last or '--' in last or 'Hidden' in last or '[' in last:
                continue
            if not first or '--' in first or 'Living' in first or '[' in first:
                continue
                if to_len(first) < 1 or to_len(last) < 1:
                    continue
            if (to_pos_int(dyr) < 1850 and to_pos_int(byr) < 19250) or to_pos_int(byr) < 1800:
                logger.warning(
                    f"need birth place(s) for {last} {first}, bpl:[{bpl}] fbpl:[{fbpl}] mbpl[{mbpl}]   byr:{byr} - dyr:{dyr} : 0")
                continue
            lspan = to_pos_int(dyr) - to_pos_int(byr)
            if (lspan < 21) or to_pos_int(byr) > 1935:
                continue

            row = {}
            if i.get('sex', '').upper() == 'M':
                row.update({
                    'h_first': first, 'h_last': last, 'h_byr': byr, 'h_bmo': bmo,
                    'h_bpl': bpl, 'h_fbpl': fbpl, 'h_mbpl': mbpl
                })
                for yr in range(1850, 1960, 10):
                    row[f'h_rsst_{yr}'] = res[str(yr)]
                    row[f'h_rscty_{yr}'] = res_cty[str(yr)]
                row.update({
                    'h_dyr': dyr, 'h_dpl': dpl,
                    'w_first': '', 'w_last': '', 'w_byr': '', 'w_bmo': '',
                    'w_bpl': '', 'w_fbpl': '', 'w_mbpl': ''
                })
                for yr in range(1850, 1960, 10):
                    row[f'w_rsst_{yr}'] = ''
                    row[f'w_rscty_{yr}'] = ''
                row.update({
                    'w_dyr': '', 'w_dpl': '',
                    'marr_yr': '', 'marr_pl': '', 'num_children': 0, 'kid_fingerprint': 0
                })

                writeSQL(first, last, byr, bpl, mbpl, fbpl)

            else:
                row.update({
                    'h_first': '', 'h_last': '', 'h_byr': '', 'h_bmo': '',
                    'h_bpl': '', 'h_fbpl': '', 'h_mbpl': ''
                })
                for yr in range(1850, 1960, 10):
                    row[f'h_rsst_{yr}'] = ''
                    row[f'h_rscty_{yr}'] = ''
                row.update({
                    'h_dyr': '', 'h_dpl': '',
                    'w_first': first, 'w_last': last, 'w_byr': byr, 'w_bmo': bmo,
                    'w_bpl': bpl, 'w_fbpl': fbpl, 'w_mbpl': mbpl
                })
                for yr in range(1850, 1960, 10):
                    row[f'w_rsst_{yr}'] = res[str(yr)]
                    row[f'w_rscty_{yr}'] = res_cty[str(yr)]
                row.update({
                    'w_dyr': dyr, 'w_dpl': dpl,
                    'marr_yr': '', 'marr_pl': '', 'num_children': 0, 'kid_fingerprint': 0
                })

            records.append(row)

    with open(r'../../JSON/gedcom_couples.json', 'w', encoding='utf-8') as f:
        json.dump(records, f, indent=4, ensure_ascii=False)
    return records


def write_csv(rows, filepath):
    if not rows:
        return
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    logger.info(f"  CSV written: {filepath}")


def write_xlsx(indi_rows, fam_rows, couple_rows, filepath):
    wb = Workbook()

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', start_color='2F5496')
    alt_fill = PatternFill('solid', start_color='EBF3FB')
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    normal_font = Font(name='Arial', size=9)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def style_header_row(ws, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.row_dimensions[1].height = 28

    def write_sheet(ws, rows, title):
        if not rows:
            return
        cols = list(rows[0].keys())
        ws.title = title
        ws.append(cols)
        style_header_row(ws, len(cols))
        for i, row in enumerate(rows, start=2):
            ws.append([row.get(c, '') for c in cols])
            fill = alt_fill if i % 2 == 0 else None
            for col_idx in range(1, len(cols) + 1):
                cell = ws.cell(row=i, column=col_idx)
                cell.font = normal_font
                cell.alignment = wrap_align
                if fill:
                    cell.fill = fill
        for col_idx, col_name in enumerate(cols, start=1):
            max_len = len(col_name)
            for row in rows:
                val = str(row.get(col_name, ''))
                seg = max((s.strip() for s in val.split('|')), key=len, default='')
                max_len = max(max_len, len(seg))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)
        ws.freeze_panes = 'A2'

    ws1 = wb.active
    write_sheet(ws1, indi_rows, 'Individuals')

    ws2 = wb.create_sheet('Families')
    write_sheet(ws2, fam_rows, 'Families')

    ws3 = wb.create_sheet('Couples')
    write_sheet(ws3, couple_rows, 'Couples')

    ws4 = wb.create_sheet('Summary')
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 20
    ws4['A1'].value = 'GEDCOM Export Summary'
    ws4['A1'].font = Font(name='Arial', bold=True, size=14)
    ws4['A1'].alignment = Alignment(horizontal='center')
    ws4.merge_cells('A1:B1')

    stats = [
        ('Total Individuals', len(indi_rows)),
        ('Total Families', len(fam_rows)),
        ('Total Couples', len(couple_rows)),
        ('With Birth Date', sum(1 for r in indi_rows if r['Birth Date'])),
        ('With Death Date', sum(1 for r in indi_rows if r['Death Date'])),
        ('With Burial Info', sum(1 for r in indi_rows if r['Burial Date'] or r['Burial Place'])),
        ('With Marriage', sum(1 for r in fam_rows if r['Marriage Date'])),
        ('Males', sum(1 for r in indi_rows if r['Sex'] == 'M')),
        ('Females', sum(1 for r in indi_rows if r['Sex'] == 'F')),
    ]
    for row_num, (label, value) in enumerate(stats, start=3):
        ws4.cell(row=row_num, column=1, value=label).font = Font(name='Arial', bold=True, size=10)
        c = ws4.cell(row=row_num, column=2, value=value)
        c.font = Font(name='Arial', size=10)
        c.alignment = Alignment(horizontal='right')

    wb.save(filepath)
    logger.info(f"  XLSX written: {filepath}")


if __name__ == '__main__':
    places = []

    logger = gen_logging.setup_logging(logger_name="gedcom_analysis")

    parser = argparse.ArgumentParser(description="Parse GEDCOM and export to CSV and formatted XLSX.")
    parser.add_argument("-i", "--input", default=INPUT_GEDCOM, help="Path to input GEDCOM file.")
    parser.add_argument("-o", "--outdir", default=OUTPUT_DIR, help="Path to output directory.")
    args = parser.parse_args()

    gedcom_path = args.input
    out_dir = args.outdir

    if not os.path.exists(gedcom_path):
        logger.error(f"ERROR: GEDCOM file not found: {gedcom_path}")
        sys.exit(1)

    os.makedirs(out_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(gedcom_path))[0]

    logger.info(f"Parsing: {gedcom_path}")
    individuals, families = parse_gedcom(gedcom_path)
    logger.info(f"  Found {len(individuals):,} individuals, {len(families):,} families.")

    logger.info("Analyzing Homeland Counties...")
    analyze_homeland_counties(individuals, families, logger, out_dir)

    logger.info("Building rows...")
    indi_rows = build_individuals_rows(individuals, families)
    fam_rows = build_families_rows(individuals, families)
    couple_rows = build_couples_rows(individuals, families)

    logger.info("Writing CSV files...")
    write_csv(indi_rows, os.path.join(out_dir, f"{base_name}_individuals.csv"))
    write_csv(fam_rows, os.path.join(out_dir, f"{base_name}_families.csv"))
    write_csv(couple_rows, os.path.join(out_dir, f"{base_name}_couples.csv"))

    logger.info("Writing Excel workbook...")
    write_xlsx(indi_rows, fam_rows, couple_rows, os.path.join(out_dir, f"{base_name}.xlsx"))

    places.sort()
    uplaces = set(places)

    out_f = os.path.join(out_dir, "statecodes.txt")
    with open(out_f, "a") as file:
        for p in uplaces:
            if p and len(p.strip()) > 0:
                dbg = common_utils.get_bpl_prefixes(p, "gedcom state")
                file.write(f"{dbg} - {p}\n")

    str = common_utils.dict_to_str(individuals, 1)
    logger.info(str)

    logger.info("\nDone!")
